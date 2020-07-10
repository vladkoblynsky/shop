import datetime

from django.utils.text import slugify
from measurement.measures import Weight
from openpyxl import load_workbook, Workbook
from prices import Money
from transliterate import translit

from main.product.models import Product, ProductType, ProductVariant, Category, Stock


def load_products_to_db():
    product_type = ProductType.objects.all().first()
    wb = load_workbook('./products.xlsx', data_only=True)
    activeSheet = wb.active
    max_rows = activeSheet.max_row + 1
    name_col = 'B' # A, B, C...
    variant_col = 'C'
    variant_price_col = 'H'
    variant_cost_price_col = 'G'
    unit_col = 'D'
    stock_col = 'E'
    category_col = 'I'
    prev_product = None
    for i in range(2, max_rows):
        try:
            product = prev_product
            product_name = activeSheet['%s%i' % (name_col, i)].value
            product_unit = activeSheet['%s%i' % (unit_col, i)].value
            variant_name = activeSheet['%s%i' % (variant_col, i)].value
            variant_price = activeSheet['%s%i' % (variant_price_col, i)].value
            variant_cost_price = activeSheet['%s%i' % (variant_cost_price_col, i)].value
            category_name = activeSheet['%s%i' % (category_col, i)].value
            stock_count = activeSheet['%s%i' % (stock_col, i)].value
            category_slug = slugify(category_name)
            if not category_slug:
                category_slug = slugify(translit(category_name, reversed=True))
            category, created = Category.objects.get_or_create(name=category_name, slug=category_slug)
            if not prev_product or (prev_product and (prev_product.name.strip() != product_name.strip())):
                product_slug = slugify(product_name)
                if not product_slug:
                    product_slug = slugify(translit(product_name, reversed=True))
                product = Product.objects.create(name=product_name.strip(),
                                                 publication_date=datetime.datetime.now(),
                                                 slug=product_slug,
                                                 is_published=True,
                                                 unit=product_unit,
                                                 description=product_name,
                                                 category=category,
                                                 product_type=product_type)
                prev_product = product
            if variant_name:
                variant_slug = slugify(variant_name)
                if not variant_slug:
                    variant_slug = slugify(translit(variant_name, reversed=True))
                price = Money(variant_price, "BYN")
                cost_price = Money(variant_cost_price, "BYN")
                variant_weight = Weight(kg=1.5)
                variant = ProductVariant.objects.create(sku='%s-%i' % (variant_slug, i),
                                                        name=variant_name.strip(),
                                                        price_override=price,
                                                        cost_price=cost_price,
                                                        product=product,
                                                        weight=variant_weight
                                                        )
                Stock.objects.get_or_create(quantity=stock_count, product_variant=variant)

        except Exception as e:
            print(e)
    update_product_minimal_and_maximal_price()


def update_product_minimal_and_maximal_price():
    products = Product.objects.all()
    for product in products:
        price_range = product.get_price_range()
        minimal = price_range.start.amount
        maximal = price_range.stop.amount
        product.minimal_variant_price = Money(minimal, 'BYN')
        product.maximal_variant_price = Money(maximal, 'BYN')
        product.save()